"""
excel_builder.py — Africa Fuel Tracker · Excel Workbook Generator
7 sheets: Dashboard · Raw Data · Monthly Prices · By Region · Rankings · FX Rates · Metadata
Auto-updated daily by the scheduler.
"""
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── Palette ────────────────────────────────────────────────────────────────────
DARK  = "0F3460"
NAVY  = "16213E"
RED   = "E94560"
GREEN = "00A878"
BLUE  = "1A73E8"
AMBER = "F59E0B"
CYAN  = "06B6D4"
WHITE = "FFFFFF"
ROW1  = "EAF4FB"
ROW2  = "F5F5F5"
BORD  = "CBD5E1"
SUB   = "2563EB"
TEXT  = "1E293B"

THIN = Border(
    left=Side(style="thin", color=BORD), right=Side(style="thin", color=BORD),
    top=Side(style="thin", color=BORD),  bottom=Side(style="thin", color=BORD),
)

def F(c): return PatternFill("solid", fgColor=c)
def hfont(c=WHITE, sz=10, bold=True): return Font(name="Calibri", bold=bold, size=sz, color=c)
def dfont(c=TEXT,  sz=9,  bold=False): return Font(name="Calibri", bold=bold, size=sz, color=c)
def center(): return Alignment(horizontal="center", vertical="center")
def left():   return Alignment(horizontal="left",   vertical="center")
def right():  return Alignment(horizontal="right",  vertical="center")

REG_COLORS = {
    "North Africa":   "059669",
    "West Africa":    "2563EB",
    "East Africa":    "D97706",
    "Central Africa": "7C3AED",
    "Southern Africa":"DC2626",
}


def _title(ws, text, c1, c2, row, height=30, bg=DARK, fg=WHITE, sz=13, wrap=False):
    s, e = get_column_letter(c1), get_column_letter(c2)
    ws.merge_cells(f"{s}{row}:{e}{row}")
    c = ws[f"{s}{row}"]
    c.value = text
    c.font  = Font(name="Calibri", bold=True, size=sz, color=fg)
    c.fill  = F(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    ws.row_dimensions[row].height = height


def _hdr(ws, col, row, text, bg=SUB):
    c = ws[f"{get_column_letter(col)}{row}"]
    c.value = text; c.font = hfont(); c.fill = F(bg)
    c.border = THIN; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30


def _cell(ws, col, row, val, fill=ROW1, bold=False, num=None, align="center", color=TEXT):
    c = ws[f"{get_column_letter(col)}{row}"]
    c.value = val; c.font = dfont(color, bold=bold)
    c.fill = F(fill); c.border = THIN
    c.alignment = Alignment(horizontal=align, vertical="center")
    if num: c.number_format = num
    return c


def build(records, fx_rates, cb_sources, months, out_path):
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_dashboard(wb, records, months)
    _sheet_raw(wb, records, months)
    _sheet_monthly(wb, records, months)
    _sheet_region(wb, records)
    _sheet_rankings(wb, records)
    _sheet_fx(wb, fx_rates, records, cb_sources)
    _sheet_meta(wb, months)
    wb.save(out_path)
    print(f"   ✅  Excel → {out_path}")


# ════════════════════════════════════════════════════════════════════════
def _sheet_dashboard(wb, records, months):
    ws = wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    now = datetime.datetime.utcnow()
    avg = sum(r["gas_usd"] for r in records) / len(records)
    high = max(records, key=lambda r: r["gas_usd"])
    low  = min(records, key=lambda r: r["gas_usd"])
    surge= max(records, key=lambda r: r["chg_gas"])

    _title(ws, "🌍  AFRICA REAL-TIME FUEL PRICE TRACKER  —  Jan–Mar 2026",
           2, 18, 2, height=42, sz=18)
    _title(ws,
        f"Updated: {now.strftime('%d %B %Y  |  %H:%M UTC')}  ·  54 Countries  ·  5 Regions  ·  "
        f"Prices in USD/L and Local Currency/L  ·  Sources: GlobalPetrolPrices · OPEC · World Bank",
        2, 18, 3, height=18, bg=RED, sz=9)

    kpis = [
        ("Highest Gasoline", f"${high['gas_usd']:.3f}/L", high["name"], RED),
        ("Lowest Gasoline",  f"${low['gas_usd']:.3f}/L",  low["name"],  GREEN),
        ("Africa Average",   f"${avg:.3f}/L",             "54 nations", BLUE),
        ("Biggest Jan→Mar Surge",  f"+{surge['chg_gas']:.1f}%",    surge["name"], AMBER),
        ("Countries Tracked", "54", "5 African regions", "155724"),
    ]
    for i,(lbl,val,sub,col) in enumerate(kpis):
        c1 = 2+i*3; c2=c1+2
        _title(ws, lbl, c1, c2, 5, height=18, bg=col, sz=9)
        _title(ws, val, c1, c2, 6, height=22, bg="F0F8FF", fg=col, sz=14)
        _title(ws, sub, c1, c2, 7, height=14, bg="F0F8FF", fg="666666", sz=8)

    _title(ws, "  📋  ALL COUNTRIES — MARCH 2026 LATEST PRICES  (USD + Local Currency)",
           2, 18, 9, height=22, bg=NAVY, sz=11)

    hdr_texts = [
        "#","Country","Region","Currency","FX Rate\n(LC/USD)",
        "Gas\nUSD/L","Gas\nLocal/L","Die\nUSD/L","Die\nLocal/L",
        "LPG\nUSD/kg","LPG\nLocal/kg","Jan→Mar\nGas %",
        "Jan→Mar\nDie %","vs Africa\nAvg USD","Price\nLevel","Octane\nRON",
        "Updated",
    ]
    hdr_widths=[4,22,18,9,11,12,12,11,11,11,11,10,10,11,10,8,20]
    for i,(h,w) in enumerate(zip(hdr_texts,hdr_widths)):
        ws.column_dimensions[get_column_letter(i+2)].width = w
        _hdr(ws, i+2, 10, h)

    for ri, r in enumerate(records):
        row = ri + 11
        fill = ROW1 if ri%2==0 else ROW2
        diff = round(r["gas_usd"] - sum(x["gas_usd"] for x in records)/len(records), 4)
        lev  = ("🔴 HIGH" if r["gas_usd"]>1.3 else ("🟡 MID" if r["gas_usd"]>0.8 else "🟢 LOW"))

        vals = [ri+1, r["name"], r["region"], r["currency"], r["fx_rate"],
                r["gas_usd"], r["gas_loc"], r["die_usd"], r["die_loc"],
                r["lpg_usd"], r["lpg_loc"],
                r["chg_gas"]/100, r["chg_die"]/100,
                diff, lev, r["octane"], r["updated"]]
        nfmts=[None,None,None,None,"#,##0.0000",
               "$#,##0.000","#,##0.00","$#,##0.000","#,##0.00",
               "$#,##0.000","#,##0.00",
               "+0.00%;-0.00%;0.00%","+0.00%;-0.00%;0.00%",
               "+$#,##0.000;-$#,##0.000",None,None,"@"]
        aligns=["center","left","left","center","right",
                "center","center","center","center","center","center",
                "center","center","center","center","center","center"]
        for ci,(v,nf,aln) in enumerate(zip(vals,nfmts,aligns)):
            col_color = TEXT
            if ci == 11:
                col_color = RED if r["chg_gas"]>2 else (GREEN if r["chg_gas"]<-0.5 else "555555")
            c = _cell(ws, ci+2, row, v, fill, num=nf, align=aln, color=col_color)
        ws.row_dimensions[row].height = 15

    ws.freeze_panes = "C11"
    ws.auto_filter.ref = f"B10:R{10+len(records)}"


# ════════════════════════════════════════════════════════════════════════
def _sheet_raw(wb, records, months):
    ws = wb.create_sheet("📁 Raw Data")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    _title(ws, "RAW DATA — ALL 54 COUNTRIES · PERIOD: 1 JAN 2026 → PRESENT", 2, 21, 1, height=28, sz=12)

    hdrs=[("Country",22),("Region",18),("Currency",9),("FX Rate",11),
          ("Gas Jan\nUSD",11),("Gas Jan\nLocal",12),("Gas Feb\nUSD",11),("Gas Feb\nLocal",12),("Gas Mar\nUSD",11),("Gas Mar\nLocal",12),
          ("Die Jan\nUSD",11),("Die Jan\nLocal",12),("Die Feb\nUSD",11),("Die Feb\nLocal",12),("Die Mar\nUSD",11),("Die Mar\nLocal",12),
          ("LPG Mar\nUSD",11),("LPG Mar\nLocal",12),
          ("Gas Chg\nJan→Mar",10),("Die Chg\nJan→Mar",10),("Octane",8)]
    for i,(h,w) in enumerate(hdrs):
        ws.column_dimensions[get_column_letter(i+2)].width=w
        _hdr(ws, i+2, 2, h)

    for ri, r in enumerate(records):
        row=ri+3; fill=ROW1 if ri%2==0 else ROW2
        g=r["gas_usd_w"]; d=r["die_usd_w"]; gl=r["gas_loc_w"]; dl=r["die_loc_w"]
        vals=[r["name"],r["region"],r["currency"],r["fx_rate"],
              g[0],gl[0],g[1],gl[1],g[2],gl[2],
              d[0],dl[0],d[1],dl[1],d[2],dl[2],
              r["lpg_usd"],r["lpg_loc"],r["chg_gas"]/100,r["chg_die"]/100,r["octane"]]
        nfmts=[None,None,None,"#,##0.0000",
               "$#,##0.000","#,##0.00","$#,##0.000","#,##0.00","$#,##0.000","#,##0.00",
               "$#,##0.000","#,##0.00","$#,##0.000","#,##0.00","$#,##0.000","#,##0.00",
               "$#,##0.000","#,##0.00","+0.0%;-0.0%;0.0%","+0.0%;-0.0%;0.0%",None]
        for ci,(v,nf) in enumerate(zip(vals,nfmts)):
            aln="left" if ci<2 else "center"
            col_c = TEXT
            if ci in [18,19]:
                pct = r["chg_gas"] if ci==18 else r["chg_die"]
                col_c = RED if pct>2 else (GREEN if pct<-0.5 else TEXT)
            _cell(ws,ci+2,row,v,fill,num=nf,align=aln,color=col_c)
        ws.row_dimensions[row].height = 14

    last=len(records)+2
    ws.conditional_formatting.add(f"E3:I{last}",
        ColorScaleRule(start_type="min",start_color="63BE7B",
                       mid_type="percentile",mid_value=50,mid_color="FFEB84",
                       end_type="max",end_color="F8696B"))
    ws.freeze_panes="C3"
    ws.auto_filter.ref=f"B2:V{last}"


# ════════════════════════════════════════════════════════════════════════
def _sheet_monthly(wb, records, months):
    """Weekly price timeline: 01 Jan → 20 Mar 2026 (12-13 weekly snapshots)."""
    import data as D
    ws = wb.create_sheet("📅 Weekly Timeline")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    week_labels = D.WEEK_LABELS
    n_weeks     = D.N_WEEKS

    _title(ws,
        f"WEEKLY PRICE TIMELINE — 01 JAN 2026 → 20 MAR 2026  ({n_weeks} weekly snapshots · USD/L)",
        2, n_weeks+3, 1, height=28, sz=12)

    # Column headers: Country | Region | Currency | FX | Week1 | Week2 | ...
    ws.column_dimensions[get_column_letter(2)].width = 24  # Country
    ws.column_dimensions[get_column_letter(3)].width = 18  # Region
    ws.column_dimensions[get_column_letter(4)].width = 9   # Currency
    ws.column_dimensions[get_column_letter(5)].width = 10  # FX
    ws.column_dimensions[get_column_letter(6)].width = 10  # Jan01 Loc
    for wi in range(n_weeks):
        ws.column_dimensions[get_column_letter(wi+7)].width = 10

    for ci, h in enumerate(["Country","Region","Currency","FX Rate","Jan 01 Local"]):
        _hdr(ws, ci+2, 2, h)
    for wi, wl in enumerate(week_labels):
        _hdr(ws, wi+7, 2, wl)
    ws.row_dimensions[2].height = 30

    # Data rows — gasoline USD/L weekly series
    for ri, r in enumerate(records):
        row  = ri + 3
        fill = ROW1 if ri%2==0 else ROW2
        fx   = r["fx_rate"]
        jan_loc = round(r["gas_usd_w"][0] * fx, 2)

        for ci, (v, nf, aln) in enumerate([
            (r["name"],   None,         "left"),
            (r["region"], None,         "left"),
            (r["currency"],None,        "center"),
            (fx,          "#,##0.0000", "right"),
            (jan_loc,     "#,##0.00",   "center"),
        ]):
            _cell(ws, ci+2, row, v, fill, num=nf, align=aln)

        # Weekly gas USD values
        for wi, price in enumerate(r["gas_usd_w"]):
            c = _cell(ws, wi+7, row, price, fill, num="$#,##0.000", align="center")
        ws.row_dimensions[row].height = 14

    last_row = len(records) + 2
    ws.freeze_panes = "G3"

    # Colour scale on weekly data
    first_week_col = get_column_letter(7)
    last_week_col  = get_column_letter(6 + n_weeks)
    ws.conditional_formatting.add(
        f"{first_week_col}3:{last_week_col}{last_row}",
        ColorScaleRule(start_type="min", start_color="63BE7B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="F8696B")
    )

    # Line chart — Africa average weekly trend
    chart_row = last_row + 3
    ws[f"B{chart_row}"] = "Week"
    ws[f"C{chart_row}"] = "Africa Avg Gas (USD/L)"
    ws[f"D{chart_row}"] = "Africa Avg Die (USD/L)"
    for wi, wl in enumerate(week_labels):
        ws[f"B{chart_row+wi+1}"] = wl
        avg_g = sum(r["gas_usd_w"][wi] for r in records) / len(records)
        avg_d = sum(r["die_usd_w"][wi] for r in records) / len(records)
        ws[f"C{chart_row+wi+1}"] = round(avg_g, 4)
        ws[f"D{chart_row+wi+1}"] = round(avg_d, 4)

    lc = LineChart()
    lc.title = f"Africa Average Fuel Price — 01 Jan to 20 Mar 2026 (Weekly, USD/L)"
    lc.y_axis.title = "Price (USD/L)"
    lc.x_axis.title = "Week"
    lc.style = 10; lc.width = 28; lc.height = 14; lc.smooth = True
    lc.add_data(
        Reference(ws, min_col=3, max_col=4, min_row=chart_row, max_row=chart_row+n_weeks),
        titles_from_data=True
    )
    lc.set_categories(Reference(ws, min_col=2, min_row=chart_row+1, max_row=chart_row+n_weeks))
    ws.add_chart(lc, f"B{chart_row+n_weeks+3}")


# ════════════════════════════════════════════════════════════════════════
def _sheet_region(wb, records):
    ws = wb.create_sheet("🗺️ By Region")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width=2
    _title(ws,"PRICES BY REGION — MARCH 2026 LATEST  (USD/L and Local Currency/L)",2,12,1,height=28,sz=12)

    REGIONS=["North Africa","West Africa","East Africa","Central Africa","Southern Africa"]
    row=3; chart_data_start=None

    for region in REGIONS:
        recs=[r for r in records if r["region"]==region]
        col=REG_COLORS[region]
        _title(ws,f"  🌐  {region.upper()}  ({len(recs)} countries)",2,12,row,height=22,bg=col,sz=10)
        row+=1
        for ci,(h,w) in enumerate([("Country",22),("Currency",9),("FX Rate",11),
                                     ("Gas USD/L",12),("Gas Local/L",13),("Die USD/L",11),
                                     ("Die Local/L",12),("LPG USD/kg",11),("LPG Local/kg",12),
                                     ("Jan→Mar %",10),("vs Avg USD",11)]):
            ws.column_dimensions[get_column_letter(ci+2)].width=w
            _hdr(ws,ci+2,row,h)
        row+=1
        avg_reg=sum(r["gas_usd"] for r in recs)/len(recs)
        for ri,r in enumerate(recs):
            fill=ROW1 if ri%2==0 else ROW2
            diff=round(r["gas_usd"]-avg_reg,4)
            vals=[r["name"],r["currency"],r["fx_rate"],r["gas_usd"],r["gas_loc"],
                  r["die_usd"],r["die_loc"],r["lpg_usd"],r["lpg_loc"],r["chg_gas"]/100,diff]
            nfmts=[None,None,"#,##0.0000","$#,##0.000","#,##0.00","$#,##0.000","#,##0.00",
                   "$#,##0.000","#,##0.00","+0.0%;-0.0%;0.0%","+$#,##0.000;-$#,##0.000"]
            for ci2,(v,nf) in enumerate(zip(vals,nfmts)):
                aln="left" if ci2==0 else "center"
                col_c=TEXT
                if ci2==9: col_c=RED if r["chg_gas"]>2 else (GREEN if r["chg_gas"]<-0.5 else TEXT)
                if ci2==10: col_c=RED if diff>0 else GREEN
                _cell(ws,ci2+2,row,v,fill,num=nf,align=aln,color=col_c)
            ws.row_dimensions[row].height=14; row+=1
        avg_g=sum(r["gas_usd"] for r in recs)/len(recs)
        avg_d=sum(r["die_usd"] for r in recs)/len(recs)
        _title(ws,f"  Avg {region}: Gas ${avg_g:.3f}/L · Die ${avg_d:.3f}/L",
               2,12,row,height=16,bg=col,fg=WHITE,sz=9); row+=3

    # Bar chart
    crow=row+1
    if chart_data_start is None: chart_data_start=crow
    ws[f"B{crow}"]="Region"; ws[f"C{crow}"]="Gas USD/L"; ws[f"D{crow}"]="Die USD/L"
    for ri2,region in enumerate(REGIONS):
        recs2=[r for r in records if r["region"]==region]
        ws[f"B{crow+ri2+1}"]=region
        ws[f"C{crow+ri2+1}"]=round(sum(r["gas_usd"] for r in recs2)/len(recs2),3)
        ws[f"D{crow+ri2+1}"]=round(sum(r["die_usd"] for r in recs2)/len(recs2),3)
    bc=BarChart(); bc.type="col"; bc.grouping="clustered"
    bc.title="Regional Avg Gasoline & Diesel (USD/L) — March 2026"
    bc.y_axis.title="Price (USD/L)"; bc.style=10; bc.width=24; bc.height=14
    bc.add_data(Reference(ws,min_col=3,max_col=4,min_row=crow,max_row=crow+5),titles_from_data=True)
    bc.set_categories(Reference(ws,min_col=2,min_row=crow+1,max_row=crow+5))
    ws.add_chart(bc,f"B{crow+7}")
    ws.freeze_panes="B3"


# ════════════════════════════════════════════════════════════════════════
def _sheet_rankings(wb, records):
    ws = wb.create_sheet("🏆 Rankings")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width=2
    _title(ws,"FUEL PRICE RANKINGS — ALL 54 AFRICAN COUNTRIES (March 2026 · USD/L)",2,17,1,height=28,sz=12)

    sorted_gas=sorted(records, key=lambda r: r["gas_usd"], reverse=True)
    for group_title,group,bg,col_off in [
        ("🔴 TOP 10 MOST EXPENSIVE",sorted_gas[:10],RED,2),
        ("🟢 TOP 10 MOST AFFORDABLE",sorted_gas[-10:][::-1],GREEN,10),
    ]:
        _title(ws,group_title,col_off,col_off+6,3,height=20,bg=bg,sz=10)
        for ci,(h,w) in enumerate([("Rank",5),("Country",22),("Region",16),
                                    ("Gas USD/L",13),("Gas Local/L",14),("Currency",9),("Jan→Mar %",10)]):
            ws.column_dimensions[get_column_letter(col_off+ci)].width=w
            _hdr(ws,col_off+ci,4,h)
        for ri,r in enumerate(group):
            row=5+ri; fill=ROW1 if ri%2==0 else ROW2
            col_c=RED if bg==RED else GREEN
            for ci2,(v,nf) in enumerate(zip(
                [ri+1,r["name"],r["region"],r["gas_usd"],r["gas_loc"],r["currency"],r["chg_gas"]/100],
                [None,None,None,"$#,##0.000","#,##0.00",None,"+0.0%;-0.0%;0.0%"]
            )):
                aln="left" if ci2==1 else "center"
                c2=_cell(ws,col_off+ci2,row,v,fill,num=nf,align=aln)
                if ci2==3: c2.font=dfont(col_c,bold=True)
            ws.row_dimensions[row].height=14

    # Full horizontal bar chart
    crow=18
    ws[f"B{crow}"]="Country"; ws[f"C{crow}"]="Gas USD/L"
    for i,r in enumerate(sorted_gas):
        ws[f"B{crow+i+1}"]=r["name"]; ws[f"C{crow+i+1}"]=r["gas_usd"]
    bch=BarChart(); bch.type="bar"
    bch.title="All 54 Countries — Gasoline Ranking (USD/L)"
    bch.style=10; bch.width=22; bch.height=52
    bch.add_data(Reference(ws,min_col=3,max_col=3,min_row=crow,max_row=crow+54),titles_from_data=True)
    bch.set_categories(Reference(ws,min_col=2,min_row=crow+1,max_row=crow+54))
    ws.add_chart(bch,"P3")
    ws.conditional_formatting.add(f"C{crow+1}:C{crow+54}",
        ColorScaleRule(start_type="min",start_color="63BE7B",
                       mid_type="percentile",mid_value=50,mid_color="FFEB84",
                       end_type="max",end_color="F8696B"))


# ════════════════════════════════════════════════════════════════════════
def _sheet_fx(wb, fx_rates, records, cb_sources):
    ws = wb.create_sheet("💱 FX Rates")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width=2
    for i,w in enumerate([16,34,16,16,40]):
        ws.column_dimensions[get_column_letter(i+2)].width=w
    _title(ws,"EXCHANGE RATES — LOCAL CURRENCY vs USD  (Reference Rates · March 2026)",2,6,1,height=28,sz=12)
    for ci,h in enumerate(["Currency Code","Country / Zone","LC per 1 USD","USD per 1 LC","Central Bank / Source"]):
        _hdr(ws,ci+2,2,h)

    from collections import defaultdict
    cur_map=defaultdict(list)
    for r in records: cur_map[r["currency"]].append(r["name"])

    for ri,(code,rate) in enumerate(sorted(fx_rates.items())):
        row=ri+3; fill=ROW1 if ri%2==0 else ROW2
        zone=" / ".join(cur_map.get(code,[code]))[:50]
        usd_per=round(1/rate,8) if rate else 0
        for ci,(v,nf,aln) in enumerate([
            (code,None,"center"),(zone,None,"left"),
            (rate,"#,##0.0000","right"),(usd_per,"0.000000","right"),
            (cb_sources.get(code,"—"),None,"left")
        ]):
            _cell(ws,ci+2,row,v,fill,num=nf,align=aln,
                  color=CYAN if ci==0 else TEXT)
        ws.row_dimensions[row].height=16

    n=len(fx_rates)+4
    ws.merge_cells(f"B{n}:F{n}")
    c=ws[f"B{n}"]
    c.value=("ℹ️  All rates are indicative reference rates as of March 2026. "
             "XOF (BCEAO) and XAF (BEAC) are pegged to EUR at 655.957 CFA/EUR (≈ 603.5 CFA/USD). "
             "For live rates the pipeline fetches from open.er-api.com on each run.")
    c.font=Font(name="Calibri",size=9,italic=True,color="555555")
    c.alignment=Alignment(wrap_text=True,vertical="top")
    ws.row_dimensions[n].height=36


# ════════════════════════════════════════════════════════════════════════
def _sheet_meta(wb, months):
    ws = wb.create_sheet("ℹ️ Metadata")
    ws.sheet_view.showGridLines = False
    for i,w in enumerate([3,32,62]): ws.column_dimensions[get_column_letter(i+1)].width=w
    _title(ws,"METADATA · SOURCES · METHODOLOGY",2,3,1,height=30,sz=13)
    now=datetime.datetime.utcnow()
    rows=[
        ("GENERAL",""),
        ("Period","1 January 2026 → present (updated daily by scheduler)"),
        ("Last Updated",now.strftime("%d %B %Y — %H:%M UTC")),
        ("Countries","54 African nations (all AU member states)"),
        ("Regions","North Africa (7) · West Africa (15) · East Africa (14) · Central Africa (8) · Southern Africa (10)"),
        ("",""),("DATA SOURCES",""),
        ("Primary","GlobalPetrolPrices.com — weekly retail pump prices (all taxes/subsidies inclusive)"),
        ("Secondary","OPEC Annual Statistical Bulletin"),
        ("Tertiary","World Bank Commodity Price Data (Pink Sheet)"),
        ("FX Rates","National Central Banks · OANDA · XE.com — fetched live via open.er-api.com on each run"),
        ("",""),("METHODOLOGY",""),
        ("Price Basis","Retail pump price including all taxes, levies, subsidies — final consumer price"),
        ("USD Conversion","Local price ÷ FX rate (LC/USD)"),
        ("Local Price","USD price × FX rate (rounded to 2 dp)"),
        ("Change","% change Jan 2026 → latest month"),
        ("Africa Average","Simple arithmetic mean of all 54 country gasoline USD prices"),
        ("",""),("AUTOMATION",""),
        ("Scheduler","scheduler.py — runs daily at configured time (default 06:00)"),
        ("Pipeline","fetch FX → build records → write Excel → write HTML → git push → GitHub Pages"),
        ("Deployment","GitHub Pages — same URL always; updates ~30s after push"),
    ]
    SECTIONS={"GENERAL","DATA SOURCES","METHODOLOGY","AUTOMATION"}
    for ri,(k,v) in enumerate(rows):
        r=ri+2; is_sec=k in SECTIONS
        bg=NAVY if is_sec else (ROW1 if ri%2==0 else ROW2)
        for ci,val in enumerate([k,v]):
            col=get_column_letter(ci+2)
            c=ws[f"{col}{r}"]
            c.value=val; c.border=THIN
            c.font=Font(name="Calibri",bold=is_sec,size=10 if is_sec else 9,
                        color=WHITE if is_sec else TEXT)
            c.fill=F(bg)
            c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
        ws.row_dimensions[r].height=20
