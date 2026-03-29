"""
generate_excel.py — Africa Fuel Tracker  |  Professional Excel Report
Reads  data/prices_db.json + data/history_db.json
Writes africa_fuel_tracker_YYYY-MM-DD.xlsx

Sheets
------
1. Summary       — KPI cards, regional overview, data quality
2. All Countries — 54 countries, conditional formatting, data bars
3. Price History — Jan 2026 → today, USD + local, change icons
4. Rankings      — Top/Bottom 10, biggest movers
5. By Region     — one block per region with regional averages
"""
import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

ROOT       = Path(__file__).parent
DATA_DIR   = ROOT / "data"
PRICES_DB  = DATA_DIR / "prices_db.json"
HISTORY_DB = DATA_DIR / "history_db.json"

REGIONS = [
    "North Africa", "West Africa", "Central Africa",
    "East Africa",  "Southern Africa",
]
REG_DARK = {
    "North Africa":    "7C4D00",
    "West Africa":     "065F46",
    "Central Africa":  "7C2D12",
    "East Africa":     "1E3A8A",
    "Southern Africa": "7F1D1D",
}
REG_LIGHT = {
    "North Africa":    "FEF9EE",
    "West Africa":     "F0FDF4",
    "Central Africa":  "FFF7ED",
    "East Africa":     "EFF6FF",
    "Southern Africa": "FFF5F5",
}
REG_MID = {
    "North Africa":    "FDE68A",
    "West Africa":     "A7F3D0",
    "Central Africa":  "FDBA74",
    "East Africa":     "BFDBFE",
    "Southern Africa": "FECACA",
}

def fill(h):    return PatternFill("solid", fgColor=h)
def fnt(bold=False, size=10, color="111827", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")
def aln(h="left", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)
def brd(color="D1D5DB", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, row, col, value=None, *, bold=False, size=10, color="111827",
       bg=None, h="left", v="center", wrap=False, indent=0,
       italic=False, nf=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = fnt(bold=bold, size=size, color=color, italic=italic)
    c.alignment = aln(h=h, v=v, wrap=wrap, indent=indent)
    if bg:     c.fill          = fill(bg)
    if nf:     c.number_format = nf
    if border: c.border        = border
    return c

def mg(ws, r, c1, c2, value=None, *, bold=False, size=11, color="111827",
       bg=None, h="center", v="center", italic=False, indent=0):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(row=r, column=c1, value=value)
    c.font      = fnt(bold=bold, size=size, color=color, italic=italic)
    c.alignment = aln(h=h, v=v, indent=indent)
    if bg: c.fill = fill(bg)
    return c

def rh(ws, r, h):  ws.row_dimensions[r].height = h
def cw(ws, c, w):  ws.column_dimensions[get_column_letter(c)].width = w

def chg(v):
    if   v >  0.1: return f"▲ +{v:.1f}%", "991B1B", "FEE2E2"
    elif v < -0.1: return f"▼ {v:.1f}%",  "166534", "DCFCE7"
    else:          return "—",             "6B7280", "F9FAFB"

def conf(c):
    return {"high":"🟢 High","medium":"🟡 Medium","low":"🔴 Low"}.get(c, c)

def load_data():
    p    = json.loads(PRICES_DB.read_text(encoding="utf-8"))
    h    = json.loads(HISTORY_DB.read_text(encoding="utf-8"))
    meta = p["meta"]
    rows = []
    for country, d in p["data"].items():
        hist  = h.get(country, [])
        first = next((e for e in hist if e["date"] >= "2026-01-01"),
                     hist[0] if hist else None)
        gf    = first["gas_usd"] if first else d["gas_usd"]
        df    = first["die_usd"] if first else d["die_usd"]
        glf   = first["gas_loc"] if first else d["gas_loc"]
        dlf   = first["die_loc"] if first else d["die_loc"]
        cg    = round((d["gas_usd"] - gf) / gf * 100, 2) if gf else 0
        cd    = round((d["die_usd"] - df) / df * 100, 2) if df else 0
        rows.append({
            "country": country,       "iso2": d["iso2"],
            "region":  d["region"],   "currency": d["currency"],
            "fx_rate": d["fx_rate"],  "gas_usd": d["gas_usd"],
            "die_usd": d["die_usd"],  "gas_loc": d["gas_loc"],
            "die_loc": d["die_loc"],  "gas_usd_jan": gf,
            "die_usd_jan": df,        "gas_loc_jan": glf,
            "die_loc_jan": dlf,       "chg_gas": cg,
            "chg_die": cd,            "confidence": d["confidence"],
            "effective_date": d["effective_date"],
            "old_source": d.get("old_source", False),
            "stale": d.get("stale", False),
            "source_url": d.get("source_url", ""),
        })
    rows.sort(key=lambda x: (
        REGIONS.index(x["region"]) if x["region"] in REGIONS else 9,
        x["country"],
    ))
    return meta, rows

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
def sh_summary(wb, meta, rows):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "065F46"
    for i, w in enumerate([2,28,16,16,16,16,16,2], 1): cw(ws, i, w)

    for r2 in [1,2,3]:
        for c2 in range(1,9): ws.cell(row=r2,column=c2).fill = fill("0F172A")
    rh(ws,1,8); rh(ws,2,44); rh(ws,3,20)
    mg(ws,2,2,7, "⛽  AFRICA FUEL TRACKER  ·  2026",
       bold=True,size=24,color="F8FAFC",bg="0F172A",h="left",indent=1)
    mg(ws,3,2,7,
       f"Official fuel price intelligence  ·  54 African nations  ·  Updated: {meta.get('run_date','')}",
       size=10,color="94A3B8",bg="0F172A",h="left",indent=1)

    gas_v  = [r["gas_usd"] for r in rows]
    die_v  = [r["die_usd"] for r in rows]
    max_r  = max(rows, key=lambda x: x["gas_usd"])
    min_r  = min(rows, key=lambda x: x["gas_usd"])
    mvrs   = [r for r in rows if abs(r["chg_gas"]) > 0.1]
    big    = max(mvrs, key=lambda x: x["chg_gas"]) if mvrs else None
    kpis   = [
        (2,"54","countries tracked","TOTAL","065F46","D1FAE5"),
        (3,f"${sum(gas_v)/len(gas_v):.3f}","Africa avg gas USD/L","AVG GAS","1D4ED8","DBEAFE"),
        (4,f"${max_r['gas_usd']:.3f}",max_r["country"],"HIGHEST","991B1B","FEE2E2"),
        (5,f"${min_r['gas_usd']:.3f}",min_r["country"],"LOWEST","166534","DCFCE7"),
        (6,f"+{big['chg_gas']:.1f}%" if big else "—",
             big["country"] if big else "No change","BIGGEST Δ","7C3AED","EDE9FE"),
        (7,f"${sum(die_v)/len(die_v):.3f}","Africa avg diesel USD/L","AVG DIESEL","0E7490","CFFAFE"),
    ]
    rh(ws,4,10); rh(ws,5,16); rh(ws,6,38); rh(ws,7,18); rh(ws,8,10)
    for col,val,sub,lbl,acc,bg in kpis:
        ws.cell(row=5,column=col,value=lbl).font = fnt(bold=True,size=8,color=acc)
        ws.cell(row=5,column=col).fill           = fill(bg)
        ws.cell(row=5,column=col).alignment      = aln(h="center")
        ws.cell(row=5,column=col).border         = Border(
            left=Side(style="thin",color="E5E7EB"),
            right=Side(style="thin",color="E5E7EB"),
            top=Side(style="thin",color="E5E7EB"))
        v = ws.cell(row=6,column=col,value=val)
        v.font      = Font(bold=True,size=20,color=acc,name="Calibri")
        v.fill      = fill(bg); v.alignment = aln(h="center")
        v.border    = Border(left=Side(style="thin",color="E5E7EB"),
                             right=Side(style="thin",color="E5E7EB"))
        s = ws.cell(row=7,column=col,value=sub)
        s.font      = fnt(size=9,color="6B7280")
        s.fill      = fill(bg); s.alignment = aln(h="center")
        s.border    = Border(left=Side(style="thin",color="E5E7EB"),
                             right=Side(style="thin",color="E5E7EB"),
                             bottom=Side(style="medium",color=acc))

    rh(ws,9,10); rh(ws,10,26)
    mg(ws,10,2,7,"REGIONAL OVERVIEW",bold=True,size=11,
       color="F8FAFC",bg="1E293B",h="left",indent=1)
    rh(ws,11,20)
    for i,h2 in enumerate(["Region","Countries","Avg Gas (USD/L)","Avg Diesel (USD/L)","Min Gas","Max Gas"],2):
        c2=sc(ws,11,i,h2,bold=True,size=9,color="F8FAFC",bg="334155",h="center")
        c2.border=Border(bottom=Side(style="medium",color="4B6584"))
    for idx,reg in enumerate(REGIONS):
        rr2=12+idx; rg=[r for r in rows if r["region"]==reg]
        gv2=[r["gas_usd"] for r in rg]; dv2=[r["die_usd"] for r in rg]
        bg2="F8FAFC" if idx%2==0 else "FFFFFF"; b2=brd("E2E8F0")
        sc(ws,rr2,2,f"  {reg}",bold=True,size=10,color=REG_DARK.get(reg,"374151"),bg=bg2,border=b2)
        sc(ws,rr2,3,len(rg),size=10,color="374151",bg=bg2,h="center",border=b2)
        sc(ws,rr2,4,sum(gv2)/len(gv2) if gv2 else 0,size=10,color="374151",
           bg=bg2,h="right",nf="$#,##0.000",border=b2)
        sc(ws,rr2,5,sum(dv2)/len(dv2) if dv2 else 0,size=10,color="374151",
           bg=bg2,h="right",nf="$#,##0.000",border=b2)
        sc(ws,rr2,6,min(gv2) if gv2 else 0,size=10,color="166534",
           bg=bg2,h="right",nf="$#,##0.000",border=b2)
        sc(ws,rr2,7,max(gv2) if gv2 else 0,size=10,color="991B1B",
           bg=bg2,h="right",nf="$#,##0.000",border=b2)
        rh(ws,rr2,20)

    rh(ws,18,10); rh(ws,19,26)
    mg(ws,19,2,7,"DATA QUALITY",bold=True,size=11,color="F8FAFC",bg="1E293B",h="left",indent=1)
    dq=[
        (2,"🟢 High confidence",  sum(1 for r in rows if r["confidence"]=="high"),   "166534","DCFCE7"),
        (3,"🟡 Medium confidence",sum(1 for r in rows if r["confidence"]=="medium"), "92400E","FEF3C7"),
        (4,"🔴 Low confidence",   sum(1 for r in rows if r["confidence"]=="low"),    "991B1B","FEE2E2"),
        (5,"⚠️ Stale prices",     sum(1 for r in rows if r["stale"]),               "B45309","FFFBEB"),
        (6,"🕰️ Old source",       sum(1 for r in rows if r["old_source"]),          "7C3AED","EDE9FE"),
        (7,"📈 Prices changed",   sum(1 for r in rows if abs(r["chg_gas"])>0.1),    "1D4ED8","EFF6FF"),
    ]
    rh(ws,20,18); rh(ws,21,32)
    for col,lbl,val,clr,bg in dq:
        sc(ws,20,col,lbl,size=9,color="374151",bg=bg,h="center",border=brd("E2E8F0"))
        c2=ws.cell(row=21,column=col,value=val)
        c2.font=Font(bold=True,size=18,color=clr,name="Calibri")
        c2.fill=fill(bg); c2.alignment=aln(h="center")
        c2.border=Border(left=Side(style="thin",color="E5E7EB"),
                         right=Side(style="thin",color="E5E7EB"),
                         bottom=Side(style="medium",color=clr))
    rh(ws,23,8); rh(ws,24,16)
    mg(ws,24,2,7,
       "Sources: Official national energy regulators  ·  FX: Frankfurter API (ECB) & central bank fixed rates",
       italic=True,size=8,color="9CA3AF",bg="FFFFFF",h="left",indent=1)

# ── Sheet 2: All Countries ────────────────────────────────────────────────────
def sh_all(wb, meta, rows):
    ws = wb.create_sheet("All Countries")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "1D4ED8"
    for i,w in enumerate([22,17,7,9,13,13,14,14,13,11,13],1): cw(ws,i,w)

    for c2 in range(1,12): ws.cell(row=1,column=c2).fill = fill("0F172A")
    mg(ws,1,1,11,f"⛽  ALL COUNTRIES — CURRENT FUEL PRICES  ·  {meta.get('run_date','')}",
       bold=True,size=13,color="F1F5F9",bg="0F172A",h="left",indent=1)
    rh(ws,1,32)

    H=2
    hdrs=["Country","Region","ISO","Currency",
          "Gas (USD/L)","Die (USD/L)","Gas (Local/L)","Die (Local/L)",
          "FX Rate","Chg Jan→Now","Confidence"]
    for i,h2 in enumerate(hdrs,1):
        c2=sc(ws,H,i,h2,bold=True,size=9,color="F1F5F9",bg="1E3A5F",h="center")
        c2.border=Border(bottom=Side(style="medium",color="3B82F6"),
                         right=Side(style="thin",color="2D5A8E"))
    rh(ws,H,24)

    prev=None; r=H+1; DS=r
    for row in rows:
        if row["region"]!=prev:
            prev=row["region"]
            dk=REG_DARK.get(row["region"],"374151")
            lk=REG_LIGHT.get(row["region"],"F3F4F6")
            for c2 in range(1,12): ws.cell(row=r,column=c2).fill=fill(lk)
            mg(ws,r,1,11,f"  ▸  {row['region'].upper()}",
               bold=True,size=9,color=dk,bg=lk,h="left",indent=1)
            ws.row_dimensions[r].height=16; r+=1

        bg0="F8FAFC" if r%2==0 else "FFFFFF"
        b2=brd("E2E8F0")
        ct,cfg,cbg=chg(row["chg_gas"])
        vals=[
            (row["country"],   True, "111827",bg0,  "left",  None),
            (row["region"],    False,"6B7280", bg0,  "left",  None),
            (row["iso2"],      False,"9CA3AF", bg0,  "center",None),
            (row["currency"],  True, "1D4ED8", bg0,  "center",None),
            (row["gas_usd"],   False,"111827", bg0,  "right", "$#,##0.000"),
            (row["die_usd"],   False,"111827", bg0,  "right", "$#,##0.000"),
            (row["gas_loc"],   False,"374151", bg0,  "right", "#,##0.00"),
            (row["die_loc"],   False,"374151", bg0,  "right", "#,##0.00"),
            (row["fx_rate"],   False,"9CA3AF", bg0,  "right", "#,##0.00"),
            (ct,               True, cfg,      cbg,  "center",None),
            (conf(row["confidence"]),False,"374151",bg0,"center",None),
        ]
        for ci,(v,bd,fg,bg,ha,nf) in enumerate(vals,1):
            c2=ws.cell(row=r,column=ci,value=v)
            c2.font=Font(name="Calibri",size=10,bold=bd,color=fg)
            c2.fill=fill(bg); c2.alignment=Alignment(horizontal=ha,vertical="center")
            c2.border=b2
            if nf: c2.number_format=nf
        ws.row_dimensions[r].height=18; r+=1

    DE=r-1
    ws.conditional_formatting.add(f"E{DS}:E{DE}",
        ColorScaleRule(start_type="min",start_color="4ADE80",
                       mid_type="percentile",mid_value=50,mid_color="FDE047",
                       end_type="max",end_color="F87171"))
    ws.conditional_formatting.add(f"F{DS}:F{DE}",
        ColorScaleRule(start_type="min",start_color="4ADE80",
                       mid_type="percentile",mid_value=50,mid_color="FDE047",
                       end_type="max",end_color="F87171"))
    ws.conditional_formatting.add(f"E{DS}:E{DE}",
        DataBarRule(start_type="min",end_type="max",color="3B82F6",showValue=True))
    ws.freeze_panes="A3"
    r+=1
    mg(ws,r,1,4,"AFRICA AVERAGE  (54 nations)",bold=True,size=9,
       color="F1F5F9",bg="1E3A5F",h="left",indent=1)
    sc(ws,r,5,sum(x["gas_usd"] for x in rows)/len(rows),
       bold=True,size=10,color="F1F5F9",bg="1E3A5F",h="right",nf="$#,##0.000")
    sc(ws,r,6,sum(x["die_usd"] for x in rows)/len(rows),
       bold=True,size=10,color="F1F5F9",bg="1E3A5F",h="right",nf="$#,##0.000")
    for c2 in range(7,12): ws.cell(row=r,column=c2).fill=fill("1E3A5F")
    ws.row_dimensions[r].height=20

# ── Sheet 3: Price History ────────────────────────────────────────────────────
def sh_history(wb, meta, rows):
    ws = wb.create_sheet("Price History")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "166534"
    for i,w in enumerate([22,8,13,13,12,13,13,12,13,13,12],1): cw(ws,i,w)

    for c2 in range(1,12): ws.cell(row=1,column=c2).fill=fill("0F172A")
    mg(ws,1,1,11,f"⛽  PRICE HISTORY  ·  JAN 2026 → {meta.get('run_date','')}",
       bold=True,size=13,color="F1F5F9",bg="0F172A",h="left",indent=1)
    rh(ws,1,32)

    rh(ws,2,18)
    for c2 in range(1,12): ws.cell(row=2,column=c2).fill=fill("1E293B")
    mg(ws,2,3,5,"⛽  GASOLINE (USD/L)",bold=True,size=9,color="FCD34D",bg="1E293B")
    mg(ws,2,6,8,"🚛  DIESEL (USD/L)",bold=True,size=9,color="93C5FD",bg="1E293B")
    mg(ws,2,9,11,"💱  LOCAL CURRENCY / L",bold=True,size=9,color="86EFAC",bg="1E293B")

    H=3
    hdrs=["Country","Currency","Jan 2026","Latest","Change",
          "Jan 2026","Latest","Change","Gas Local","Die Local","Confidence"]
    hbg=["1E293B","1E293B","14532D","14532D","14532D",
         "1E3A5F","1E3A5F","1E3A5F","1A3A2E","1A3A2E","2D1B69"]
    for i,(h2,hb) in enumerate(zip(hdrs,hbg),1):
        c2=sc(ws,H,i,h2,bold=True,size=9,color="F1F5F9",bg=hb,h="center")
        c2.border=Border(bottom=Side(style="medium",color="6EE7B7"),
                         right=Side(style="thin",color="374151"))
    rh(ws,H,22)

    prev=None; r=H+1; DS=r
    for row in rows:
        if row["region"]!=prev:
            prev=row["region"]
            dk=REG_DARK.get(row["region"],"374151")
            lk=REG_LIGHT.get(row["region"],"F3F4F6")
            for c2 in range(1,12): ws.cell(row=r,column=c2).fill=fill(lk)
            mg(ws,r,1,11,f"  ▸  {row['region'].upper()}",
               bold=True,size=9,color=dk,bg=lk,h="left",indent=1)
            ws.row_dimensions[r].height=15; r+=1

        bg0="F8FAFC" if r%2==0 else "FFFFFF"; b2=brd("E2E8F0")
        gt,gfg,gbg=chg(row["chg_gas"]); dt,dfg,dbg=chg(row["chg_die"])
        vals=[
            (row["country"],     True, "111827",bg0,    "left",  None),
            (row["currency"],    True, "1D4ED8",bg0,    "center",None),
            (row["gas_usd_jan"], False,"166534","F0FDF4","right", "$#,##0.000"),
            (row["gas_usd"],     True, "111827",bg0,    "right", "$#,##0.000"),
            (gt,                 True, gfg,     gbg,    "center",None),
            (row["die_usd_jan"], False,"1D4ED8","EFF6FF","right", "$#,##0.000"),
            (row["die_usd"],     True, "111827",bg0,    "right", "$#,##0.000"),
            (dt,                 True, dfg,     dbg,    "center",None),
            (row["gas_loc"],     False,"374151",bg0,    "right", "#,##0.00"),
            (row["die_loc"],     False,"374151",bg0,    "right", "#,##0.00"),
            (conf(row["confidence"]),False,"374151",bg0,"center",None),
        ]
        for ci,(v,bd,fg,bg,ha,nf) in enumerate(vals,1):
            c2=ws.cell(row=r,column=ci,value=v)
            c2.font=Font(name="Calibri",size=10,bold=bd,color=fg)
            c2.fill=fill(bg); c2.alignment=Alignment(horizontal=ha,vertical="center")
            c2.border=b2
            if nf: c2.number_format=nf
        ws.row_dimensions[r].height=18; r+=1

    DE=r-1
    ws.conditional_formatting.add(f"D{DS}:D{DE}",
        ColorScaleRule(start_type="min",start_color="4ADE80",
                       mid_type="percentile",mid_value=50,mid_color="FDE047",
                       end_type="max",end_color="F87171"))
    ws.freeze_panes="A4"

# ── Sheet 4: Rankings ─────────────────────────────────────────────────────────
def sh_rankings(wb, meta, rows):
    ws = wb.create_sheet("Rankings")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "991B1B"
    for i,w in enumerate([22,17,13,13,12, 4, 22,17,13,13,12],1): cw(ws,i,w)

    for c2 in range(1,12): ws.cell(row=1,column=c2).fill=fill("0F172A")
    mg(ws,1,1,11,"⛽  FUEL PRICE RANKINGS  ·  Gasoline (USD/L)",
       bold=True,size=13,color="F1F5F9",bg="0F172A",h="left",indent=1)
    rh(ws,1,32); rh(ws,2,8)

    def draw(sc_col, title, tbg, data, rbgs):
        mg(ws,3,sc_col,sc_col+4,title,bold=True,size=11,
           color="FFFFFF",bg=tbg,h="left",indent=1)
        rh(ws,3,28)
        for i,h2 in enumerate(["Country","Region","Gas USD/L","Diesel USD/L","Chg Jan→Now"]):
            c2=sc(ws,4,sc_col+i,h2,bold=True,size=9,color="F9FAFB",bg="374151",h="center")
            c2.border=Border(bottom=Side(style="medium",color="6B7280"))
        rh(ws,4,20)
        for idx,row in enumerate(data):
            rr=5+idx; bg=rbgs[idx%2]
            rb="FEF3C7" if idx<3 else bg; rf="92400E" if idx<3 else "111827"
            b2=brd("E2E8F0"); ct,cfg,cbg=chg(row["chg_gas"])
            c2=ws.cell(row=rr,column=sc_col,value=f"  {idx+1}.  {row['country']}")
            c2.font=Font(name="Calibri",size=10,bold=True,color=rf)
            c2.fill=fill(rb); c2.alignment=aln(h="left"); c2.border=b2
            sc(ws,rr,sc_col+1,row["region"],size=9,color="6B7280",bg=bg,border=b2)
            sc(ws,rr,sc_col+2,row["gas_usd"],bold=True,size=11,color="111827",
               bg=bg,h="right",nf="$#,##0.000",border=b2)
            sc(ws,rr,sc_col+3,row["die_usd"],size=10,color="374151",
               bg=bg,h="right",nf="$#,##0.000",border=b2)
            sc(ws,rr,sc_col+4,ct,bold=True,size=10,color=cfg,bg=cbg,h="center",border=b2)
            ws.row_dimensions[rr].height=22

    sd=sorted(rows,key=lambda x:-x["gas_usd"])
    sa=sorted(rows,key=lambda x: x["gas_usd"])
    draw(1,"🔴  TOP 10 MOST EXPENSIVE GAS","991B1B",sd[:10],["FEF2F2","FFF5F5"])
    draw(7,"🟢  TOP 10 MOST AFFORDABLE GAS","166534",sa[:10],["F0FDF4","F7FEF9"])

    rh(ws,16,10); rh(ws,17,26)
    mg(ws,17,1,11,"📊  BIGGEST PRICE CHANGES  vs Jan 2026",
       bold=True,size=11,color="F8FAFC",bg="1E293B",h="left",indent=1)
    mvrs=sorted([r for r in rows if abs(r["chg_gas"])>0.1],key=lambda x:-x["chg_gas"])
    if not mvrs:
        rh(ws,18,20)
        mg(ws,18,1,11,"No significant price changes since Jan 2026",
           italic=True,size=10,color="6B7280")
    else:
        for i,h2 in enumerate(["Country","Region","Currency","Jan 2026","Latest","Change","Δ USD/L"],1):
            c2=sc(ws,18,i,h2,bold=True,size=9,color="F9FAFB",bg="334155",h="center")
            c2.border=Border(bottom=Side(style="medium",color="64748B"))
        rh(ws,18,20)
        for idx,row in enumerate(mvrs):
            rr=19+idx; bg="F8FAFC" if idx%2==0 else "FFFFFF"
            b2=brd("E2E8F0"); ct,cfg,cbg=chg(row["chg_gas"])
            delta=round(row["gas_usd"]-row["gas_usd_jan"],4)
            vals=[
                (row["country"],    True, "111827",bg,    "left", None),
                (row["region"],     False,"6B7280", bg,    "left", None),
                (row["currency"],   False,"1D4ED8", bg,    "center",None),
                (row["gas_usd_jan"],False,"166534","F0FDF4","right","$#,##0.000"),
                (row["gas_usd"],    True, "111827",bg,    "right","$#,##0.000"),
                (ct,                True, cfg,      cbg,   "center",None),
                (delta,             True, cfg,      cbg,   "right","$#,##0.0000"),
            ]
            for ci,(v,bd,fg,bgc,ha,nf) in enumerate(vals,1):
                c2=ws.cell(row=rr,column=ci,value=v)
                c2.font=Font(name="Calibri",size=10,bold=bd,color=fg)
                c2.fill=fill(bgc); c2.alignment=Alignment(horizontal=ha,vertical="center")
                c2.border=b2
                if nf: c2.number_format=nf
            ws.row_dimensions[rr].height=20

# ── Sheet 5: By Region ────────────────────────────────────────────────────────
def sh_region(wb, meta, rows):
    ws = wb.create_sheet("By Region")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "7C3AED"
    for i,w in enumerate([22,8,13,13,13,13,12,11],1): cw(ws,i,w)

    for c2 in range(1,9): ws.cell(row=1,column=c2).fill=fill("0F172A")
    mg(ws,1,1,8,"⛽  PRICES BY REGION  ·  Sorted by region & country",
       bold=True,size=13,color="F1F5F9",bg="0F172A",h="left",indent=1)
    rh(ws,1,32)

    r=2
    for reg in REGIONS:
        rg=[row for row in rows if row["region"]==reg]
        if not rg: continue
        dk=REG_DARK.get(reg,"374151"); lk=REG_LIGHT.get(reg,"F3F4F6")
        mk=REG_MID.get(reg,"E5E7EB")

        rh(ws,r,10); r+=1
        for c2 in range(1,9): ws.cell(row=r,column=c2).fill=fill(dk)
        mg(ws,r,1,8,f"  {reg.upper()}  ·  {len(rg)} countries",
           bold=True,size=11,color="FFFFFF",bg=dk,h="left",indent=1)
        rh(ws,r,26); r+=1

        gv=[x["gas_usd"] for x in rg]; dv=[x["die_usd"] for x in rg]
        sc(ws,r,1,"Regional average →",bold=True,size=9,color=dk,bg=mk)
        sc(ws,r,3,sum(gv)/len(gv) if gv else 0,bold=True,size=9,color=dk,
           bg=mk,h="right",nf="$#,##0.000")
        sc(ws,r,4,sum(dv)/len(dv) if dv else 0,bold=True,size=9,color=dk,
           bg=mk,h="right",nf="$#,##0.000")
        sc(ws,r,5,min(gv) if gv else 0,size=9,color="166534",
           bg=mk,h="right",nf="$#,##0.000")
        sc(ws,r,6,max(gv) if gv else 0,size=9,color="991B1B",
           bg=mk,h="right",nf="$#,##0.000")
        for c2 in [2,7,8]: ws.cell(row=r,column=c2).fill=fill(mk)
        rh(ws,r,16); r+=1

        for i,h2 in enumerate(["Country","Currency","Gas USD/L","Diesel USD/L",
                                "Gas Local","Diesel Local","Chg Jan→Now","Confidence"],1):
            c2=sc(ws,r,i,h2,bold=True,size=9,color="F9FAFB",bg="334155",h="center")
            c2.border=Border(bottom=Side(style="medium",color=dk))
        rh(ws,r,18); r+=1

        for idx,row in enumerate(rg):
            bg="F8FAFC" if idx%2==0 else "FFFFFF"; b2=brd("E2E8F0")
            ct,cfg,cbg=chg(row["chg_gas"])
            vals=[
                (row["country"],  True, "111827",bg,  "left",  None),
                (row["currency"], True, "1D4ED8",bg,  "center",None),
                (row["gas_usd"],  False,"111827",bg,  "right", "$#,##0.000"),
                (row["die_usd"],  False,"111827",bg,  "right", "$#,##0.000"),
                (row["gas_loc"],  False,"374151",bg,  "right", "#,##0.00"),
                (row["die_loc"],  False,"374151",bg,  "right", "#,##0.00"),
                (ct,              True, cfg,     cbg, "center",None),
                (conf(row["confidence"]),False,"374151",bg,"center",None),
            ]
            for ci,(v,bd,fg,bgc,ha,nf) in enumerate(vals,1):
                c2=ws.cell(row=r,column=ci,value=v)
                c2.font=Font(name="Calibri",size=10,bold=bd,color=fg)
                c2.fill=fill(bgc); c2.alignment=Alignment(horizontal=ha,vertical="center")
                c2.border=b2
                if nf: c2.number_format=nf
            ws.row_dimensions[r].height=18; r+=1

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    for path in [PRICES_DB, HISTORY_DB]:
        if not path.exists():
            print(f"❌ {path} not found"); return

    meta, rows = load_data()
    wb = Workbook()
    wb.remove(wb.active)

    sh_summary(wb,  meta, rows)
    sh_all(wb,      meta, rows)
    sh_history(wb,  meta, rows)
    sh_rankings(wb, meta, rows)
    sh_region(wb,   meta, rows)

    run_date = meta.get("run_date", date.today().isoformat())
    out = ROOT / f"africa_fuel_tracker_{run_date}.xlsx"
    wb.save(out)
    print(f"✅ {out.name}  ({out.stat().st_size//1024} KB)")
    print(f"   Sheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
